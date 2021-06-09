import openpyxl as pyxl

filename = input("Enter the filepath: ")
wb = pyxl.load_workbook(filename)

sheetName = input("Sheet name: ")
sheet = wb[sheetName].values

start = int(input("Row with headers (this is assumed to be the start of data): ")) - 1
headers = []
# for i, row in enumerate(sheet):
#     print(row)
#     if i == start:
#         headers = row

numFilters = int(input("How many filters? "))
filters = []
for i in range(numFilters):
    print("Filter", i+1)
    f = {}
    f['column'] = int(input("Column number: ")) - 1
    f["value"] = input("Value: ")
    val = input("Contains? (y/n): ")
    if val == "y":
        f["contains"] = True
    else:
        f["contains"] = False
    filters.append(f)
filtered = []
x = filters[0]
for i, row in enumerate(sheet):
    if i == start:
        headers = row
    if row[0] == None:
        break
    elif i > start:
        good = True
        for x in filters:
            col = x["column"]
            if (x["contains"] and x["value"] not in row[col]) or (not x["contains"] and x["value"] in row[col]):
                good = False
        if good:
            element = []
            for z in row:
                if z == None:
                    break
                else:
                    element.append(z)
            filtered.append(element)

newFilename = input("New file location/name: ")
newFile = pyxl.Workbook()
newSheet = newFile.create_sheet(title=sheetName)

for i, h in enumerate(headers):
    newSheet.cell(row=1, column=i+1, value=h)
if len(filtered) != 0:
    for r in range(1, len(filtered)+1):
        for c in range(len(filtered[0])):
            newSheet.cell(row=r+1, column=c+1, value=filtered[r-1][c])
newFile.save(filename=newFilename)