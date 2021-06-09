import openpyxl as pyxl

filename = input("Enter the filepath: ")
wb = pyxl.load_workbook(filename)

sheetName = input("Sheet name: ")
sheet = wb[sheetName].values

start = int(input("Row with headers (this is assumed to be the start of data): ")) - 1
headers = []
# for i, row in enumerate(sheet):
#     print(row)
#     if i == start:
#         headers = row

whitelist = input("Whitelist text file name: ")
text = []
with open(whitelist, 'r', encoding='utf-8') as txt:
    text = txt.readlines()

column = int(input("Column with data: ")) - 1

filtered = []
for i, row in enumerate(sheet):
    if i == start:
        headers = row
    if row[0] == None:
        break
    elif i > start:
        url = row[column]
        if "/" in url:
            s = url.split('/')
            url = s[2]
        good = False
        for x in text:
            val = x.strip('\n')
            if url == val:
                good = True
        if not good:
            element = []
            for z in row:
                if z == None:
                    break
                else:
                    element.append(z)
            filtered.append(element)

newFilename = input("New file location/name: ")
newFile = pyxl.Workbook()
newSheet = newFile.create_sheet(title=sheetName)

for i, h in enumerate(headers):
    newSheet.cell(row=1, column=i+1, value=h)
if len(filtered) != 0:
    for r in range(1, len(filtered)+1):
        for c in range(len(filtered[0])):
            if c > len(filtered[r-1]):
                newSheet.cell(row=r+1, column=c+1, value=None)
            else:
                newSheet.cell(row=r+1, column=c+1, value=filtered[r-1][c])
newFile.save(filename=newFilename)